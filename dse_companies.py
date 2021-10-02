from bs4 import BeautifulSoup
import requests
import time
# import xlsxwriter
import openpyxl

# workbook = openpyxl.Workbook('dse.xlsx')
# worksheet = workbook.add_worksheet()
# sheet_row = 0
# sheet_column = 0

workbook = openpyxl.Workbook()
# workbook['Sheet'].title = 'DSE companies'
sheet = workbook.active
print(workbook.sheetnames)
sheet.title = 'DSE companies'
sheet.append(['DSE companies'])


url = 'https://www.dsebd.org/company_listing.php'

page = requests.get(url)
# print(page)
soup = BeautifulSoup(page.content, 'html.parser')
companies = soup.find_all('div', class_="BodyContent")

# 'div', class_='col-lg-4 col-md-4 col-sm-6 col-xs-12 background-white')  # find('div', class_="BodyContent").find('a', class_='ab1').text

list_of_comapanies = []
for company in companies:
    classes = company.find_all('a', class_='ab1')
    for c in classes:
        # print(c.text)
        list_of_comapanies.append(c.text)


print(len(list_of_comapanies))
for idx, item in enumerate(list_of_comapanies):
    print(idx+1, " ", item)
    sheet.append([item])
    # worksheet.write(sheet_row, sheet_column, i)
    # sheet_row += 1

workbook.save('dse.xlsx')
